import io
import re
import logging
import unicodedata

from openpyxl import load_workbook

from app.models.insurance import NetworkCoverage

logger = logging.getLogger(__name__)


def parse_excel_to_markdown(content: bytes) -> str:
    """
    Converts an Excel workbook (.xlsx) to a markdown string suitable for
    passing to Gemini's structured extraction prompt.

    Each sheet becomes a section with a markdown table.
    """
    wb = load_workbook(io.BytesIO(content), data_only=True)
    sections = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        if not rows:
            continue

        non_empty_rows = [r for r in rows if any(c is not None for c in r)]
        if not non_empty_rows:
            continue

        # Determine max columns used
        max_cols = max(len(r) for r in non_empty_rows)

        def cell_str(v) -> str:
            if v is None:
                return ""
            return str(v).strip().replace("|", "/")

        header_row = non_empty_rows[0]
        headers = [cell_str(h) or f"Col{i+1}" for i, h in enumerate(header_row)]
        # Pad or trim to max_cols
        headers = (headers + [f"Col{i+1}" for i in range(len(headers), max_cols)])[:max_cols]

        md_lines = [f"## Hoja: {sheet_name}", ""]
        md_lines.append("| " + " | ".join(headers) + " |")
        md_lines.append("| " + " | ".join(["---"] * max_cols) + " |")

        for row in non_empty_rows[1:]:
            cells = [cell_str(row[i] if i < len(row) else None) for i in range(max_cols)]
            md_lines.append("| " + " | ".join(cells) + " |")

        sections.append("\n".join(md_lines))

    if not sections:
        raise ValueError("El archivo Excel no contiene datos.")

    return "\n\n".join(sections)


def _normalize(text: str) -> str:
    """Lowercase, strip accents and surrounding whitespace."""
    text = (text or "").strip().lower()
    nfkd = unicodedata.normalize("NFKD", text)
    return "".join(c for c in nfkd if not unicodedata.combining(c))


# Each NetworkCoverage field maps to keywords that may appear in the column header.
_COLUMN_KEYWORDS = {
    "ciudad": ["ciudad"],
    "red": ["red"],
    "provider": ["prestador", "proveedor", "hospital", "clinica", "centro"],
    "address": ["direccion", "ubicacion"],
    "service_category": ["tipo de atencion", "tipo atencion"],
    "copay": ["copago", "cobertura"],
    "service_name": ["servicio"],
    "specialty": ["especialidad"],
}


def _map_columns(headers: list[str]) -> dict:
    """Maps NetworkCoverage fields to column indices by header keyword."""
    mapping = {}
    for idx, header in enumerate(headers):
        hn = _normalize(header)
        for field, keywords in _COLUMN_KEYWORDS.items():
            if field not in mapping and any(kw in hn for kw in keywords):
                mapping[field] = idx
                break
    return mapping


def _cell(row, mapping: dict, field: str) -> str:
    idx = mapping.get(field)
    if idx is None or idx >= len(row):
        return ""
    value = row[idx]
    return str(value).strip() if value is not None else ""


def _parse_copay(raw: str) -> tuple[str, float]:
    """Parses a copay cell into (copay_type, copay_value).

    "$0,00 COPAGO" -> ("fixed", 0.0)   |   "% COBERTURA" -> ("percentage", 0.0)
    """
    text = _normalize(raw)
    if not text:
        return ("fixed", 0.0)
    match = re.search(r"(\d+(?:[.,]\d+)?)", text)
    value = float(match.group(1).replace(",", ".")) if match else 0.0
    if "%" in text or "cobertura" in text:
        return ("percentage", value)
    return ("fixed", value)


def parse_excel_network_directory(content: bytes) -> list[NetworkCoverage]:
    """Parses a network directory Excel into NetworkCoverage records — no LLM.

    Each non-empty data row becomes one record. Columns are matched by header
    keyword, so their order does not matter.
    """
    wb = load_workbook(io.BytesIO(content), data_only=True)
    records: list[NetworkCoverage] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = [r for r in ws.iter_rows(values_only=True) if any(c is not None for c in r)]
        if len(rows) < 2:
            continue

        headers = [str(c) if c is not None else "" for c in rows[0]]
        mapping = _map_columns(headers)

        if "provider" not in mapping or "service_name" not in mapping:
            logger.warning(
                f"Hoja '{sheet_name}': sin columnas de proveedor/servicio reconocibles, se omite"
            )
            continue

        for row in rows[1:]:
            provider = _cell(row, mapping, "provider")
            service_name = _cell(row, mapping, "service_name")
            if not provider or not service_name:
                continue

            copay_type, copay_value = _parse_copay(_cell(row, mapping, "copay"))

            specialty = _cell(row, mapping, "specialty")
            aliases = []
            if specialty and _normalize(specialty) not in ("no aplica", "n/a", "na"):
                aliases = [specialty]

            records.append(
                NetworkCoverage(
                    red=_cell(row, mapping, "red"),
                    ciudad=_cell(row, mapping, "ciudad"),
                    provider=provider,
                    address=_cell(row, mapping, "address"),
                    service_name=service_name,
                    service_category=_cell(row, mapping, "service_category"),
                    specialty_aliases=aliases,
                    copay_type=copay_type,
                    copay_value=copay_value,
                )
            )

    logger.info(f"Excel network directory parsed: {len(records)} records")
    return records
